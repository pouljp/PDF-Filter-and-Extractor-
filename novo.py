import PyPDF2
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
import re

def extrair_informacoes(conteudo):
    # (códigos anteriores permanecem iguais)

    # Função auxiliar para extrair padrões de regex
    def extrair_padrao(padrao, texto):
        correspondencia = re.search(padrao, texto, re.DOTALL)
        return correspondencia.group(1).strip() if correspondencia else None

    # Dicionário para armazenar as informações extraídas
    informacoes = {
        "Processo Nº": extrair_padrao(padrao_processo, conteudo),
        "E.P. Nº": extrair_padrao(padrao_ep, conteudo),
        "Vara": extrair_padrao(padrao_vara, conteudo),
        "Comarca": extrair_padrao(padrao_comarca, conteudo),
        "Autor": extrair_padrao(padrao_autor, conteudo),
        "Advogado": extrair_padrao(padrao_advogado, conteudo),
        "Entidade": extrair_padrao(padrao_entidade, conteudo),
        "Ação": extrair_padrao(padrao_acao, conteudo),
        "Liquidação Valor em R$": extrair_padrao(padrao_liquidação, conteudo),
        "VALOR PRINCIPAL R$": extrair_padrao(padrao_valor_principal, conteudo),
        "DESCONTO PREVIDENCIÁRIO R$": extrair_padrao(padrao_desconto_previdenciario, conteudo),  # Novo padrão
        "ASSISTÊNCIA MÉDICA R$": extrair_padrao(padrao_assistencia_medica, conteudo),  # Novo padrão
        "SUB-TOTAL R$": extrair_padrao(padrao_sub_total, conteudo),  # Novo padrão
        "JUROS MORATÓRIOS R$": extrair_padrao(padrao_juros, conteudo),
        "HONORÁRIOS ADVOCATÍCIOS R$": extrair_padrao(padrao_honorarios, conteudo),
        "CUSTAS R$": extrair_padrao(padrao_custas, conteudo),
        "EMBARGOS A EXECUÇÃO R$": extrair_padrao(padrao_embargos, conteudo),
        "DESPESAS R$": extrair_padrao(padrao_despesas, conteudo),
        "TOTAL R$": extrair_padrao(padrao_total, conteudo),
        "Banco": extrair_padrao(padrao_banco, conteudo),  # Novo padrão
        "Agência": extrair_padrao(padrao_agencia, conteudo),  # Novo padrão
        "Conta Judicial": extrair_padrao(padrao_conta_judicial, conteudo),
        "Data do Depósito": extrair_padrao(padrao_data_deposito, conteudo),
        "Valor do Depósito": extrair_padrao(padrao_valor_deposito, conteudo),
        "OBS.": extrair_padrao(padrao_obs, conteudo)
    }

    return informacoes

def aplicar_filtro(pdf_path, output_text_path, excel_path):
    with open(pdf_path, 'rb') as file:
        pdf_reader = PyPDF2.PdfReader(file)
        num_pages = len(pdf_reader.pages)

        wb = Workbook()
        ws = wb.active

        # Adicionar cabeçalhos
        ws.append(['Linha Separadora', 'Informação', 'Conteúdo'])

        conteudo_texto = []  # Lista para armazenar o conteúdo do texto

        for page_num in range(num_pages):
            page = pdf_reader.pages[page_num]
            text = page.extract_text()

            if "DEPRE 3.4" in text:
                start_index = text.find("DEPRE 3.4")
                end_index = text.rfind("DEPRE 3.4") + len("DEPRE 3.4")

                resultado_filtrado = text[start_index:end_index].replace("DEPRE 3.4", "").strip()
                linha_separadora = '-' * 50

                # Adicionar dados à planilha Excel
                ws.append([linha_separadora, f"Filtro aplicado na página {page_num + 1}:", resultado_filtrado])

                # Adicionar linha separadora ao conteúdo do texto
                conteudo_texto.append(linha_separadora)
                # Adicionar conteúdo filtrado à lista para o arquivo de texto
                conteudo_texto.append(resultado_filtrado)

        # Adicionar formatação específica
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
            for cell in row:
                cell.font = Font(size=10)

        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 100

        # Salvar planilha Excel
        wb.save(excel_path)

        # Salvar conteúdo filtrado em um arquivo de texto
        with open(output_text_path, 'w', encoding='utf-8') as txt_file:
            for conteudo in conteudo_texto:
                txt_file.write(conteudo + '\n')

        # Extrair informações específicas e imprimir
        for conteudo in conteudo_texto:
            informacoes = extrair_informacoes(conteudo)
            print("\nInformações Extraídas:")
            for chave, valor in informacoes.items():
                print(f"{chave}: {valor}")

# Substitua 'seu_arquivo.pdf', 'saida.txt' e 'saida.xlsx' pelos caminhos reais dos seus arquivos
aplicar_filtro('C:/Users/pooul/Downloads/teste2.pdf', 'C:/Users/pooul/Downloads/arquivotxt/saida.txt', 'C:/Users/pooul/Downloads/arquivotxt/saida.xlsx')
